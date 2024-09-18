import os
import pandas as pd
import requests
import openpyxl
import xlrd
from flask import current_app, jsonify, send_file, render_template, request
from flask_login import current_user
from sqlalchemy import desc
from applications.common.utils.http import fail_api, success_api
from applications.extensions import db
from applications.extensions.init_upload import photos
from applications.models import Photo, Excel
from applications.schemas import PhotoOutSchema, ExcelSchema
from applications.common.curd import model_to_dicts


def get_photo(page, limit):
    photo = Photo.query.order_by(desc(Photo.create_time)).paginate(page=page, per_page=limit, error_out=False)
    count = Photo.query.count()
    data = model_to_dicts(schema=PhotoOutSchema, data=photo.items)
    return data, count


def upload_one(photo, mime):
    filename = photos.save(photo)
    file_url = '/_uploads/photos/'+filename
    # file_url = photos.url(filename)
    upload_url = current_app.config.get("UPLOADED_PHOTOS_DEST")
    size = os.path.getsize(upload_url + '/' + filename)
    photo = Photo(name=filename, href=file_url, mime=mime, size=size)
    db.session.add(photo)
    db.session.commit()
    return file_url


def delete_photo_by_id(_id):
    photo_name = Photo.query.filter_by(id=_id).first().name
    photo = Photo.query.filter_by(id=_id).delete()
    db.session.commit()
    upload_url = current_app.config.get("UPLOADED_PHOTOS_DEST")
    os.remove(upload_url + '/' + photo_name)
    return photo


def get_excel(page, limit):
    excel = Excel.query.order_by(desc(Excel.create_time)).paginate(page=page, per_page=limit, error_out=False)
    count = Excel.query.count()
    data = model_to_dicts(schema=ExcelSchema, data=excel.items)
    return data, count


def upload_one_excel(excel, mime):
    filename = excel.filename
    file_path = os.path.join(current_app.config['UPLOADED_EXCELS_DEST'] + '/' + filename)

    excel.save(file_path)

    size = os.path.getsize(file_path)

    excel_entry = Excel(
        name=filename,
        href=file_path,
        mime=mime,
        size=size,
        repair_status='0',
        create_by=current_user.username,
        create_time=db.func.now()
    )
    db.session.add(excel_entry)
    db.session.commit()

    json = {
        'msg': '创建成功',
        'code': 0,
        'success': True,
        'data': {
            'src': file_path
        }
    }
    return jsonify(json)


def delete_excel_by_id(id):
    excel = Excel.query.filter_by(id=id).first()

    if not excel:
        return fail_api(msg='删除失败')

    upload_url = current_app.config.get("UPLOADED_EXCELS_DEST")
    file_path = os.path.join(upload_url, excel.name)

    try:
        db.session.delete(excel)
        db.session.commit()

        if os.path.exists(file_path):
            os.remove(file_path)

        return success_api(msg='删除成功')
    except Exception as e:
        db.session.rollback()
        print(e)
        return fail_api(msg='删除异常')


def download_excel_by_id(id):
    excel = Excel.query.filter_by(id=id).first()

    if not excel:
        return fail_api(msg='下载失败')

    file_name = excel.name
    file_path = os.path.join(current_app.config.get('UPLOADED_EXCELS_DEST') + '/' + file_name)

    if not os.path.isfile(file_path):
        return fail_api(msg='文件未找到')

    try:
        return send_file(file_path, as_attachment=True)

    except Exception as e:
        print(e)
        return fail_api(msg='下载异常')


def read_excel2(file_path):
    """
    读取 CSV | XLS | XLSX 文件格式的文件
    :param file_path: 文件路径
    :return: 读取成功返回 DataFrame，失败返回 None
    """
    try:
        if file_path.endswith('.csv'):
            return pd.read_csv(file_path)
        elif file_path.endswith('.xls') or file_path.endswith('.xlsx'):
            return pd.read_excel(file_path, engine='openpyxl' if file_path.endswith('.xlsx') else None)
        else:
            return None
    except Exception as e:
        print(e)
        return None


def view_excel_by_id(id):
    excel = Excel.query.filter_by(id=id).first()
    try:
        file_url = excel.href

        df = read_excel2(file_url)

        if df is None:
            return fail_api('文件类型不支持或读取失败')

        html = df.to_html(index=False, classes='table table-striped')
        return render_template('system/excel/excel_view.html', df=html, filename=excel.name)
    except Exception as e:
        print(e)
        return fail_api('查看文件异常')


def parse_excel_file(filename):
    """
    xls | xlsx | csv
    将读到的文件解析为 json
    POST 发送给地址修复 API
    回调时访问 receive 接口，将数据解析为 excel 保存至服务器

    payload = {
        "entity_no": 0,
        "entrust_id": "",
        "task_id": "gasghghweh2542",
        "begin_time": "2024-09-14 11:14:00",
        "returnUrl": "http://192.168.3.233:8000/system/api/receive",
        "customer_info": [
            {
                "customer_id": "廖化军",
                "homeAddress": "福建省厦门市思明区仙阁里114号304室福建省厦门市思明区仙阁里114号304室",
                "beginTime": "2024-09-18 09:01:35",
                "liveAddress": "福建省厦门市思明区仙阁里114号304室福建省厦门市思明区仙阁里114号304室",
                "user_identification": "421122200010104321"
            }
        ]
    }
    """
    if filename.endswith('xls'):
        workbook = xlrd.open_workbook(filename)
        sheet = workbook.sheet_by_index(0)
        column_names = [sheet.cell_value(0, col_idx) for col_idx in range(sheet.ncols)]
        for row_index in range(1, sheet.nrows):
            row_values = [sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols)]
            for column_name, value in zip(column_names, row_values):
                print(f"{column_name}: {value}")
                # switch column_names:
                #     case
            print()
    elif filename.endswith('xlsx'):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        column_names = [cell.value for cell in next(sheet.iter_rows())]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            for column_name, value in zip(column_names, row):
                print(f"{column_name}: {value}")
            print()
    elif filename.endswith('csv'):
        pass
    else:
        pass


def api_upload_excel_by_id(id):
    excel = Excel.query.filter_by(id=id).first()

    if not excel:
        return fail_api(msg='上传失败')

    file_path = excel.href

    df = read_excel2(file_path)
    print(df)
    if df is None:
        return fail_api(msg='上传失败')

    parse_excel_file(file_path)

    # data_to_send = df.to_dict(orient='records')

    headers = {
        'Content-Type': 'application/json'
    }
    try:
        response = requests.request("POST", current_app.config.get('UPLOAD_URL'), headers=headers, json=payload)
        """
        {
            "code": "200",  # 200为成功，其他为失败
            "msg": "数据接收成功"
        }
        """
        print(response)
        if 'code' in response.json():
            if response.json()['code'] != '200':
                return fail_api(msg='修复地址数据上传失败')
            return success_api(msg='修复地址数据上传成功')
        return fail_api(msg='json 解析错误')
    except Exception as e:
        print(e)
        return fail_api(msg='修复地址数据上传异常')


def api_receive_json():
    # 处理 api 返回的 json 格式数据
    pass


def download_template():
    file_path = current_app.config.get('TEMPLATE_EXCELS_DEST')
    if not os.path.isfile(file_path):
        return fail_api(msg='模板不存在')

    try:
        return send_file(file_path, as_attachment=True)

    except Exception as e:
        print(e)
        return fail_api(msg='下载失败')